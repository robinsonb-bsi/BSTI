import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import xml.etree.ElementTree as ET
import os

def save_results_to_excel():
    """
    This function loads JSON data from 'config.json' and XML data from 'plugins-2023-05-18.xml'.
    It extracts script names and risk factors from the XML data and creates an Excel workbook.
    The script then sets column widths, applies formatting to headers and cells, and writes data to the workbook.
    It also merges cells for category headers and applies grouping (collapsible) to the categories.
    Finally, it saves the workbook as 'plugin_info.xlsx'.

    Raises:
        Exception: If any error occurs during the process.
    """
    try:
        # Load JSON data
        with open('config.json') as json_file:
            json_data = json.load(json_file)

        # Load XML data
        tree = ET.parse('plugins-2023-05-18.xml')
        root = tree.getroot()

        # Get script names and risk factors
        plugin_info = {}

        for nasl_element in root.findall('.//nasl'):
            script_id_element = nasl_element.find('script_id')
            if script_id_element is not None and script_id_element.text is not None:
                script_id = script_id_element.text.strip()

                script_name_element = nasl_element.find('script_name')
                if script_name_element is not None and script_name_element.text is not None:
                    script_name = script_name_element.text.strip()

                    plugin_info[script_id] = {'script_name': script_name}

        # Create Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        # Set column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 80
        worksheet.column_dimensions['D'].width = 48
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 45

        # Apply header formatting
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(fill_type='solid', fgColor='C0C0C0')

        # Write headers
        worksheet.append(['Category', 'Plugin ID', 'Plugin Name', 'Link', 'Writeup ID', 'Writeup Name'])
        header_row = worksheet[1]
        for cell in header_row:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # Apply cell formatting
        cell_alignment = Alignment(vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment

        # Store the starting row index for each category
        category_start_rows = {}

        # Write data
        for plugin_key, plugin_value in json_data['plugins'].items():
            # Store the starting row index for the category
            category_start_rows[plugin_key] = len(worksheet['A']) + 1

            # Write category header row
            worksheet.append([plugin_key, '', '', '', '', ''])
            header_row_index = len(worksheet['A'])
            category_row = worksheet[header_row_index]
            category_row[0].fill = header_fill  # Apply fill to the first cell (Category)

            # Write scan_type and parameters in the category header row
            worksheet.cell(row=header_row_index, column=5, value=plugin_value.get('writeup_db_id', '')).fill = header_fill
            worksheet.cell(row=header_row_index, column=6, value=plugin_value.get('writeup_name', '')).fill = header_fill

            # Merge cells for the category row
            worksheet.merge_cells(start_row=header_row_index, start_column=1,
                                  end_row=header_row_index, end_column=4)

            if isinstance(plugin_value, dict) and 'ids' in plugin_value:
                for plugin_id in plugin_value['ids']:
                    script_info = plugin_info.get(str(plugin_id), {})
                    script_name = script_info.get('script_name', '')
                    link = f"https://www.tenable.com/plugins/nessus/{plugin_id}"
                    worksheet.append(['', plugin_id, script_name, link, '', ''])

        # Apply grouping (collapsible) for categories
        for category, start_row in category_start_rows.items():
            end_row = worksheet.max_row
            worksheet.row_dimensions.group(start_row, end_row)

        # Save the workbook
        workbook.save('plugin_info.xlsx')
    except Exception as e:
        print(e)
        
if __name__ == '__main__':
    try:
        print("Creating excel file")
        save_results_to_excel()
        print("Done")
    except Exception as e:
        print("An error occurred during creation: ", e)
