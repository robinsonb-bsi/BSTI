import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font
import time
import requests
from bs4 import BeautifulSoup
import os
import sys
from openpyxl.styles import PatternFill
from scripts.logging_config import log

"""
sudo sdkmanager "system-images;android-28;google_apis;x86_64"
avdmanager create avd -n Pixel_2_API_28 -k "system-images;android-28;google_apis;x86_64"
emulator -avd Pixel_2_API_28 -writable-system -no-snapshot -noaudio -no-window -gpu off
adb devices
nano /home/USERNAME/.MobSF/config.py

"""

class Mobber:
    def __init__(self, mobsf_url, scan_type, app_name):
        self.mobsf_url = mobsf_url
        self.api_base = "/api/v1"
        self.api_docs_url = f"{mobsf_url}/api_docs"
        self.pdf_url = f"{mobsf_url}{self.api_base}/download_pdf"
        self.upload_url = f"{mobsf_url}{self.api_base}/upload"
        self.scan_url = f"{mobsf_url}{self.api_base}/scan"
        self.scorecard_url = f"{mobsf_url}{self.api_base}/scorecard"
        self.scan_type = scan_type
        self.app_path = app_name
        self.api_key = self.get_api_key()
        self.headers = {"Authorization": self.api_key}
        
        
        if self.scan_type == "apk":
            if not os.path.exists(self.app_path):  # Check if the .apk file doesn't exist
                log.warning("APK does not exist, attempting to download it for you")
                
                # Remove the .apk extension if present in the app name
                if self.app_path.lower().endswith(".apk"):
                    self.app_path = self.app_path[: -len(".apk")]
                
                self.download_link = f"https://apkaio.com/download-apk/{self.app_path}"
                
                if self.download_apk():
                    # Add the .apk extension if the downloaded file doesn't have it
                    downloaded_file_name = self.app_path
                    if not downloaded_file_name.lower().endswith(".apk"):
                        downloaded_file_name += ".apk"
                    
                    # Update the app_path attribute with the downloaded file name
                    self.app_path = downloaded_file_name
                else:
                    log.error("Unable to download the APK")
                    sys.exit(1)

                
        report_name = f"{self.app_path}-mobsf-report.pdf"
        scorecard_name = f"SCORECARD-{self.app_path}-mobsf.xlsx"
        report_dir = 'reports'
        os.makedirs(report_dir, exist_ok=True)
        self.report_output_path = os.path.join(report_dir, report_name)
        self.scorecard_output_path = os.path.join(report_dir, scorecard_name)
        self.file_name, self.hash_value = self.upload_file()
        self.engine()





        
    def download_apk(self):
        log.info(f"Downloading APK from: {self.download_link}")

        response = requests.get(self.download_link)
        if response.status_code == 200:
            download_url_match = re.search(r'<a href="(https?://[^"]+)"[^>]*>click here</a>', response.text)
            if download_url_match:
                download_url = download_url_match.group(1)
                apk_response = requests.get(download_url)
                if apk_response.status_code == 200:
                    # Append .apk extension to the file path
                    apk_path = f"{self.app_path}.apk"
                    with open(apk_path, "wb") as file:
                        file.write(apk_response.content)
                    self.app_path = apk_path  # Update the app_path attribute
                    return True
                else:
                    log.error(f"Failed to download APK from {download_url}. Status code: {apk_response.status_code}")
            else:
                log.error("Download link not found in the HTML.")

        return False



    def get_api_key(self):
        response = requests.get(self.api_docs_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            api_key_element = soup.select_one("p.lead strong code")
            if api_key_element:
                api_key = api_key_element.get_text(strip=True)
                return api_key
        return None

    def upload_file(self):

        files = {"file": (self.app_path, open(self.app_path, "rb"), "application/octet-stream")}
        response = requests.post(self.upload_url, files=files, headers=self.headers)
        if response.status_code == 200:
            response_json = response.json()
            file_name = response_json["file_name"]
            hash_value = response_json["hash"]
            return file_name, hash_value
        elif response.status_code == 401:
            log.error("Authentication failed. Please check your API key.")
        else:
            log.error(f"Failed to upload file. Status code: {response.status_code}")
            log.error(f"Response: {response.text}")
        return None



    def monitor_scan(self):
        """Initiates a scan and logs its progress."""
        log.info('Scanning file...')

        data = {
            "file_name": self.file_name,
            "hash": self.hash_value,
            "scan_type": self.scan_type,
            "re_scan": "0"  # 0 for no rescan, 1 for rescan
        }

        try:
            response = requests.post(self.scan_url, headers=self.headers, data=data)
            log.info('Scan completed successfully.')
            return response
        except Exception as e:
            log.error(f'Error during scan: {e}')
            raise

    def scan_file(self):
        response = self.monitor_scan()
        if response.status_code == 200:
            return True
        else:
            log.error(f"Failed to start scan. Status code: {response.status_code}")
            log.error(f"Response: {response.text}")
            return False



    def generate_scorecard(self):
        retry = 0
        data = {
            "hash": self.hash_value
        }
        response = requests.post(self.scorecard_url, headers=self.headers, data=data)

        while response.status_code != 200:
            log.info("Waiting for the scorecard to become available...")
            time.sleep(3)
            log.warning(f"Retry number: {retry} - retrying up to 3 times...")

            if retry == 3:
                log.error("Unable to export report")
                return False
            elif response.status_code == 200:
                continue
            else:
                retry += 1

        try:
            # Convert the JSON response to a Python dictionary
            json_data = json.loads(response.content)

            # Extract the relevant information from the JSON data
            high_issues = json_data.get("high", [])
            warning_issues = json_data.get("warning", [])
            info_issues = json_data.get("info", [])
            secure_issues = json_data.get("secure", [])
            hotspot_issues = json_data.get("hotspot", [])

            # Create a new workbook
            workbook = Workbook()
            sheet = workbook.active

            # Increase the font size for the table headers
            header_font = Font(size=14, bold=True)

            # Helper function to write a header row
            def write_header_row(header_text, row_number, fill_color=None):
                sheet.append([header_text])
                header_cell = sheet.cell(row=row_number, column=1)
                header_cell.font = header_font
                if fill_color is not None:
                    header_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                sheet.merge_cells(start_row=row_number, end_row=row_number, start_column=1, end_column=3)
                
            def set_column_width(column, data):
                max_length = 0
                for row in data:
                    for cell in row:
                        if cell is not None:
                            cell_length = len(str(cell))
                            if cell_length > max_length:
                                max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2  # Adding padding and scaling factor
                sheet.column_dimensions[column].width = adjusted_width


            # Write the high issues header with a fill color
            write_header_row("High Issues", 1, "FFFF0000")  # Red color

            # Write the high issues subheaders
            sheet.append(["Title", "Description", "Section"])

            # Write the high issues to the sheet
            for issue in high_issues:
                title = issue.get("title", "")
                description = issue.get("description", "")
                section = issue.get("section", "")
                sheet.append([title, description, section])

            # Increase the column widths for the high issues table
            sheet.column_dimensions['A'].width = 80
            sheet.column_dimensions['B'].width = 160
            sheet.column_dimensions['C'].width = 10

            # Write the warning issues header with a different fill color
            write_header_row("Warning Issues", len(high_issues) + 3, "FFFFFF00")  # Yellow color

            # Write the warning issues subheaders
            sheet.append(["Title", "Description", "Section"])

            # Write the warning issues to the sheet
            for issue in warning_issues:
                title = issue.get("title", "")
                description = issue.get("description", "")
                section = issue.get("section", "")
                sheet.append([title, description, section])

            # Increase the column widths for the warning issues table
            sheet.column_dimensions['A'].width = 80
            sheet.column_dimensions['B'].width = 160
            sheet.column_dimensions['C'].width = 10

            # Write the info issues header with a different fill color
            write_header_row("Info Issues", len(high_issues) + len(warning_issues) + 5, "FF00FF00")  # Green color

            # Write the info issues subheaders
            sheet.append(["Title", "Description", "Section"])

            # Write the info issues to the sheet
            for issue in info_issues:
                title = issue.get("title", "")
                description = issue.get("description", "")
                section = issue.get("section", "")
                sheet.append([title, description, section])

            # Increase the column widths for the info issues table
            sheet.column_dimensions['A'].width = 80
            sheet.column_dimensions['B'].width = 160
            sheet.column_dimensions['C'].width = 10

            # Write the secure issues header with a different fill color
            write_header_row("Secure Issues", len(high_issues) + len(warning_issues) + len(info_issues) + 7, "FF0000FF")  # Blue color

            # Write the secure issues subheaders
            sheet.append(["Title", "Description", "Section"])

            # Write the secure issues to the sheet
            for issue in secure_issues:
                title = issue.get("title", "")
                description = issue.get("description", "")
                section = issue.get("section", "")
                sheet.append([title, description, section])

            # Increase the column widths for the secure issues table
            sheet.column_dimensions['A'].width = 80
            sheet.column_dimensions['B'].width = 160
            sheet.column_dimensions['C'].width = 10

            # Write the hotspot issues header with a different fill color
            write_header_row("Hotspot Issues", len(high_issues) + len(warning_issues) + len(info_issues) + len(secure_issues) + 9, "FFFF00FF")  # Purple color

            # Write the hotspot issues subheaders
            sheet.append(["Title", "Description", "Section"])

            # Write the hotspot issues to the sheet
            for issue in hotspot_issues:
                title = issue.get("title", "")
                description = issue.get("description", "")
                section = issue.get("section", "")
                sheet.append([title, description, section])

            # Increase the column widths for the hotspot issues table
            sheet.column_dimensions['A'].width = 80
            sheet.column_dimensions['B'].width = 160
            sheet.column_dimensions['C'].width = 10

            # Save the workbook to an Excel file
            workbook.save(self.scorecard_output_path)
            return True

        except json.JSONDecodeError as e:
            log.error("Failed to decode JSON response: %s", str(e))
            return False
        except Exception as e:
            log.error("An error occurred: %s", str(e))
            return False     
        

    def generate_report(self):
        retry = 0
        data = {
            "hash": self.hash_value
        }
        response = requests.post(self.pdf_url, headers=self.headers, data=data)
        while response.status_code != 200:
            log.info("Waiting for the report to become available...")
            time.sleep(3)
            log.warning(f"Retry number: {retry} - retrying up to 3 times...")
            
            if retry == 3:
                log.error("Unable to export report")
                return False
            elif response.status_code == 200:
                continue
            else:
                retry += 1
        
        # Save the HTML table to a file
        with open(self.report_output_path, "wb") as file:
            file.write(response.content)
        return True
    
    

    def engine(self):
        try:
            self.scan_file()
            log.success("Scan completed, exporting report ...")
            self.generate_report()
            log.success(f"Report saved to: {self.report_output_path}")
            self.generate_scorecard()
            log.success(f"Scorecard saved to: {self.scorecard_output_path}")
        except Exception as e:
            log.error(f"An error occurred: {str(e)}")
        # os.remove(self.app_path)



